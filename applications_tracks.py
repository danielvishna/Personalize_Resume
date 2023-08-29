import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import date

from openpyxl.utils import get_column_letter

def style_hyperlink(column, url, sheet):
    """
    Apply styling to the new row added to the Excel sheet.

    Args:
        sheet (Worksheet): The worksheet to which the new row has been added.
        url (str): The url link we need to use as hyperlink.
        column (int) : The index of the column we want to style.
    """
    sheet.cell(row=sheet.max_row, column=column).hyperlink = url
    sheet.cell(row=sheet.max_row, column=column).font = Font(color="000000FF", underline="single")

def style(sheet, job):
    """
    Apply styling to the new row added to the Excel sheet.

    Args:
        sheet (Worksheet): The worksheet to which the new row has been added.
        job (Job): An instance of the Job class containing job details.
    """
    # Add hyperlink to the "Link for the job ad" cell
    if hasattr(job, 'url'):
        style_hyperlink(6, job.get_url(), sheet)

    # Add hyperlink to the "Company" cell
    if hasattr(job, 'get_company_link'):
        style_hyperlink(3, job.get_company_link(), sheet)

    # Apply border and alignment styling to the cells in the new row
    new_row = sheet[sheet.max_row]
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center')  # Set alignment once
    for cell in new_row:
        cell.border = border
        cell.alignment = alignment


def create_or_append_excel(file, job):
    """
    Create or append data to an Excel workbook.

    Args:
        file (str): The path to the Excel file.
        job (Job): An instance of the Job class containing job details.
    """
    headers = ["No", "Title", "Company", "Location", "Date of send", "Link for the ad job"]

    # Check if the file exists
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
    except FileNotFoundError:
        # Create a new workbook and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        font = Font(bold=True)
        alignment = Alignment(horizontal='center')
        for cell in sheet["1:1"]:
            cell.font = font
            cell.alignment = alignment
            cell.border = border

    # Prepare data for the new row
    data = [sheet.max_row, job.get_title(), job.get_company_name(), job.get_location(),
            (date.today()).strftime("%d/%m/%Y"), job.get_title() + " " + job.get_company_name()]

    # Add data to the sheet
    sheet.append(data)

    # Apply styling to the new row
    style(sheet, job)

    # Save the workbook
    try:
        workbook.save(file)
    except PermissionError:
        print("The file is open. Please close it and try again.")
        return
    print("Added new row to the tracker")
