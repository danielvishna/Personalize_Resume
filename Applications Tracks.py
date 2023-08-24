import os

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def create_or_append_excel(file_path, headers, data):
    # Check if the file exists
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # Create a new workbook and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for cell in sheet["1:1"]:
            cell.font = Font(bold=True)

    # Add data to the sheet
    sheet.append(data)

    # Save the workbook
    workbook.save(file_path)

def get_number_of_rows(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return sheet.max_row
    except FileNotFoundError:
        return 0  # File doesn't exist, so there are 0 rows



file_path = "example.xlsx"

file_path = "example.xlsx"
num_rows = get_number_of_rows(file_path)
print(f"Number of rows in the file: {num_rows}")



if __name__ == '__main__':
    file_directory = 'C:/Users/DanielV/Documents/CV'
    file_name = 'Daniel Vishna CV.pdf'
    file_path = os.path.join(file_directory, file_name)
    headers = ["No", "Title" "Company", "Date of send", "Link for the job ad"]
    data = [str(get_number_of_rows(file_path)), "Value2", "Value3"]
    create_or_append_excel(file_path, headers, data)